"""動作確認用のサンプル xlsx を生成する。"""
import base64
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Font, Side

# 1x1 透明 PNG
_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+M8AAAMBAQDJ"
    "/pLvAAAAAElFTkSuQmCC"
)

out = Path("sample.xlsx")
img_path = Path("_tmp_sample.png")
img_path.write_bytes(_PNG)

wb = Workbook()

# --- シート1: 要件定義書ふう(結合フォーム + データ表 + 画像) ---
ws1 = wb.active
ws1.title = "機能要件"
ws1["A1"] = "機能要件一覧"
ws1.merge_cells("A1:D1")               # 見出しの横結合
ws1["A2"] = "項目"
ws1["B2"] = "内容"
ws1.merge_cells("B2:D2")
ws1["A3"] = "機能名"
ws1["B3"] = "ログイン\n(認証)"          # セル内改行
ws1.merge_cells("B3:D3")
ws1["A4"] = "優先度"
ws1["B4"] = "高"
ws1["A5"] = "作成日"
ws1["B5"] = datetime.datetime(2026, 6, 15, 10, 30, 0)
ws1["A6"] = "件数"
ws1["B6"] = "=3+4"                      # 数式(未計算 -> 式文字列フォールバック検証)
ws1["A7"] = "参照"
ws1["B7"] = "公式サイト"
ws1["B7"].hyperlink = "https://example.com/spec"
# 縦結合
ws1["A8"] = "区分"
ws1.merge_cells("A8:A9")
ws1["B8"] = "画面"
ws1["B9"] = "帳票"
ws1.add_image(XLImage(str(img_path)), "C8")  # C8 にアンカーした画像

# --- シート2: 画面設計書ふう(複数セクション/空行区切り) ---
ws2 = wb.create_sheet("画面設計 ログイン")  # シート名に空白を含める
# セクション1: 見出しバナー
ws2["B2"] = "1. ログイン画面"
ws2.merge_cells("B2:E2")
# 説明段落(全幅結合の長文)
ws2["B3"] = "本画面はユーザ認証を行う。ID とパスワードを入力し、ログインボタンを押下する。"
ws2.merge_cells("B3:E3")
# (4行目は空 → セクション区切り)
# セクション2: 小見出し + 罫線つきデータ表
ws2["B5"] = "1.1 入力項目"
ws2.merge_cells("B5:E5")
ws2["B6"] = "項目"
ws2["C6"] = "種別"
ws2["D6"] = "必須"
ws2["E6"] = "備考"
ws2["B7"] = "ID"
ws2["C7"] = "テキスト"
ws2["D7"] = "○"
ws2["E7"] = "半角英数"
ws2["B8"] = "PW"
ws2["C8"] = "パスワード"
ws2["D8"] = "○"
ws2["E8"] = "8文字以上"
# (9行目は空 → セクション区切り)
# セクション3: 図
ws2["B10"] = "1.2 画面イメージ"
ws2.merge_cells("B10:E10")
ws2.add_image(XLImage(str(img_path)), "B11")

# --- シート3: 罫線つきの表(内部に空スペーサー列 D)。表粉砕の回帰テスト ---
ws3 = wb.create_sheet("詳細設計")
ws3["B2"] = "2. 詳細設計"
ws3.merge_cells("B2:E2")
ws3["B2"].font = Font(bold=True, size=14)
# 表 B4:E7 (D 列はヘッダもデータも空だが罫線あり = スペーサー)
ws3["B4"] = "項目"
ws3["C4"] = "値"
ws3["E4"] = "単位"
for c in ("B4", "C4", "E4"):
    ws3[c].font = Font(bold=True)
ws3["B5"], ws3["C5"], ws3["E5"] = "最大同時接続数", 100, "件"
ws3["B6"], ws3["C6"], ws3["E6"] = "タイムアウト", 30, "秒"
ws3["B7"], ws3["C7"], ws3["E7"] = "リトライ回数", 3, "回"
thin = Side(style="thin")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
for r in range(4, 8):
    for col in range(2, 6):  # B..E(空の D 含む)
        ws3.cell(row=r, column=col).border = box

wb.save(out)
print("generated", out)
