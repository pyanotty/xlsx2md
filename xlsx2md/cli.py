"""xlsx2md のコマンドラインインターフェース。"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

from .converter import Options, convert
from .debug import dump_segments
from .extract import extract_sheet
from .pipeline import convert_structured
from .segment import segment


def _run_segment(xlsx_path: str, options: Options) -> int:
    """②分割の可視化モード。Markdown は出力せず、分割結果をダンプする。"""
    wb_data = load_workbook(xlsx_path, data_only=True)
    wb_form = (load_workbook(xlsx_path, data_only=False)
               if options.formula_fallback else None)
    for name in wb_data.sheetnames:
        ws = wb_data[name]
        wsf = wb_form[name] if wb_form is not None else None
        sheet = extract_sheet(ws, wsf, options)
        if sheet is None:
            print(f"# シート: {name}\n  (空のシート)\n")
            continue
        print(dump_segments(sheet, segment(sheet)))
    return 0


def _force_utf8_console() -> None:
    """Windows の旧コンソール(cp932)で日本語 print が落ちないようにする。"""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")   # Python 3.7+
        except Exception:  # noqa: BLE001
            pass


def main(argv: list[str] | None = None) -> int:
    _force_utf8_console()
    parser = argparse.ArgumentParser(
        prog="xlsx2md",
        description="Excel 文書(要件定義書・画面設計書)を Markdown に変換する。"
                    "シートごとに独立した .md を出力する。",
    )
    parser.add_argument("xlsx_path", help="変換する .xlsx ファイル")
    parser.add_argument("-o", "--out-dir", default="output",
                        help="出力先ディレクトリ (既定: output)")
    parser.add_argument("--image-dirname", default="images",
                        help="画像出力先のサブディレクトリ名 (既定: images)")
    parser.add_argument("--no-hidden", action="store_true",
                        help="隠し行/列を出力に含めない")
    parser.add_argument("--no-formula-fallback", action="store_true",
                        help="数式キャッシュが無い時に式文字列へフォールバックしない")
    parser.add_argument("--segment", action="store_true",
                        help="②分割の可視化のみ行う(Markdown は出力しない)")
    parser.add_argument("--faithful", action="store_true",
                        help="v1 方式(1シート=1HTMLテーブル)で忠実出力する")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="詳細ログを出力する")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(levelname)s: %(message)s",
    )

    if not Path(args.xlsx_path).is_file():
        print(f"エラー: ファイルが見つかりません: {args.xlsx_path}", file=sys.stderr)
        return 1

    options = Options(
        image_dirname=args.image_dirname,
        include_hidden=not args.no_hidden,
        formula_fallback=not args.no_formula_fallback,
    )

    if args.segment:
        return _run_segment(args.xlsx_path, options)

    if args.faithful:
        produced = convert(args.xlsx_path, args.out_dir, options)
    else:
        produced = convert_structured(args.xlsx_path, args.out_dir, options)

    print(f"{len(produced)} 個の Markdown を生成しました:")
    for p in produced:
        print(f"  {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
