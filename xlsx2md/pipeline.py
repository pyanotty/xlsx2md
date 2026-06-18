"""構造復元パイプライン: ①抽出 → ②分割 → ③役割付与 → ⑤生成 → 検証。

「設計書として自然な Markdown」を出力する v2 のメイン経路。
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from .converter import Options, _extract_images, _sanitize
from .extract import extract_sheet
from .interpret import Diagram, interpret
from .mermaid import build_mermaid
from .render import render_document
from .segment import segment
from .shapes import extract_diagrams
from .verify import verify_content

logger = logging.getLogger("xlsx2md")


def convert_structured(xlsx_path: str | Path, out_dir: str | Path,
                       options: Options | None = None) -> list[Path]:
    """xlsx を構造復元して自然な Markdown を出力し、生成パス一覧を返す。"""
    options = options or Options()
    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)

    wb_data = load_workbook(xlsx_path, data_only=True)
    wb_form = (load_workbook(xlsx_path, data_only=False)
               if options.formula_fallback else None)
    diagrams_by_sheet = extract_diagrams(xlsx_path)   # 画面遷移図など描画オブジェクト

    doc_dir = out_dir / _sanitize(xlsx_path.stem)
    images_dir = doc_dir / options.image_dirname
    doc_dir.mkdir(parents=True, exist_ok=True)

    produced: list[Path] = []
    for idx, name in enumerate(wb_data.sheetnames, start=1):
        ws = wb_data[name]
        wsf = wb_form[name] if wb_form is not None else None
        diagrams = diagrams_by_sheet.get(name, [])

        image_map = _extract_images(ws, images_dir, _sanitize(name), options)
        sheet = extract_sheet(ws, wsf, options)
        if sheet is None and not diagrams:
            md = f"# {name}\n\n_(空のシート)_\n"
        elif sheet is None:
            # 図のみのシート
            blocks = [Diagram(build_mermaid(d)) for d in _sorted(diagrams)]
            md = render_document(_DiagramOnlySheet(name), blocks,
                                 options.image_dirname, image_map)
        else:
            regions = segment(sheet)
            cell_blocks = interpret(sheet, regions, image_map)
            blocks = _merge_in_order(regions, cell_blocks, diagrams)
            md = render_document(sheet, blocks, options.image_dirname, image_map)
            _report_missing(sheet, md)
            _report_diagrams(name, diagrams)

        out_path = doc_dir / f"{idx:02d}_{_sanitize(name)}.md"
        out_path.write_text(md, encoding="utf-8")
        produced.append(out_path)
        logger.info("wrote %s", out_path)

    return produced


class _DiagramOnlySheet:
    """図しか無いシート用の最小シム(render_document はタイトルしか参照しない)。"""
    def __init__(self, name: str):
        self.name = name


def _sorted(diagrams):
    return sorted(diagrams, key=lambda d: (d.anchor_row, d.anchor_col))


def _merge_in_order(regions, cell_blocks, diagrams):
    """セル由来ブロックと図を、行位置(アンカー)で読み順に併合する。

    同じ行ではセル側を先に置く(キーの第2要素 0=セル, 1=図)。
    """
    items = [(regions[i].r0, 0, blk) for i, blk in enumerate(cell_blocks)]
    for d in diagrams:
        items.append((d.anchor_row, 1, Diagram(build_mermaid(d))))
    items.sort(key=lambda x: (x[0], x[1]))
    return [blk for _, _, blk in items]


def _report_diagrams(name: str, diagrams) -> None:
    for d in diagrams:
        logger.info("シート '%s': 図を検出(ノード %d, エッジ %d%s)",
                    name, len(d.nodes), len(d.edges),
                    f", 明示接続なしで除外した矢印 {d.skipped_connectors}"
                    if d.skipped_connectors else "")


def _report_missing(sheet, md: str) -> None:
    missing = verify_content(sheet, md)
    if missing:
        logger.warning("シート '%s': 出力に現れない元テキストが %d 件あります:",
                       sheet.name, len(missing))
        for m in missing:
            logger.warning("  - %s", m)
