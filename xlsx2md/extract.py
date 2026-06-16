"""① 抽出層: Worksheet を解釈しやすい SheetModel に変換する。

v1(converter.py)の値整形・アンカー検出を再利用しつつ、
② 以降の段が必要とする「結合・書式・画像セル」を構造化して保持する。
出力(Markdown)には出さない書式も、③ の役割判定の手がかりとして読む。
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

from .converter import _anchor_cell, _format_value


@dataclass
class CellInfo:
    """セル(結合の左上=origin)の書式サマリ。③ の役割判定に使う。"""
    bold: bool = False
    font_size: float | None = None
    has_fill: bool = False
    has_border: bool = False
    h_align: str | None = None


@dataclass
class SheetModel:
    """1 シートの構造化表現。"""
    name: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    # origin(結合左上 or 通常セル)の非空テキストのみ保持
    origin_text: dict[tuple[int, int], str] = field(default_factory=dict)
    span_of: dict[tuple[int, int], tuple[int, int]] = field(default_factory=dict)
    covered_to_origin: dict[tuple[int, int], tuple[int, int]] = field(default_factory=dict)
    style: dict[tuple[int, int], CellInfo] = field(default_factory=dict)
    image_cells: dict[tuple[int, int], int] = field(default_factory=dict)
    # 罫線のある物理セル(空セルも含む)。表領域の保護に使う
    bordered: set[tuple[int, int]] = field(default_factory=set)
    # 罫線が「表の構造」か「方眼紙全体の装飾」かの判定。装飾なら分割に使わない
    borders_are_structural: bool = False
    # 本文の代表フォントサイズ(見出し判定の基準)
    body_font_size: float | None = None

    # --- 問い合わせ用ヘルパ ---
    def origin_of(self, r: int, c: int) -> tuple[int, int]:
        return self.covered_to_origin.get((r, c), (r, c))

    def text_at(self, r: int, c: int) -> str:
        return self.origin_text.get(self.origin_of(r, c), "")

    def has_image(self, r: int, c: int) -> bool:
        return self.image_cells.get((r, c), 0) > 0

    def is_ruled(self, r: int, c: int) -> bool:
        """罫線で囲まれた表セルか(空セルでも True)。装飾罫線は無視。"""
        if not self.borders_are_structural:
            return False
        return (r, c) in self.bordered or self.origin_of(r, c) in self.bordered

    def is_blank(self, r: int, c: int) -> bool:
        # 罫線つきの空セルも「表の一部」として非空扱い → XY-cut で粉砕させない
        return (self.text_at(r, c) == ""
                and not self.has_image(r, c)
                and not self.is_ruled(r, c))

    def addr(self, r: int, c: int) -> str:
        return f"{get_column_letter(c)}{r}"


def extract_sheet(ws, wsf, options) -> SheetModel | None:
    """Worksheet から SheetModel を構築する。内容が無ければ None。"""
    # 結合情報
    span_of: dict[tuple[int, int], tuple[int, int]] = {}
    covered_to_origin: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        top = (mr.min_row, mr.min_col)
        span_of[top] = (mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != top:
                    covered_to_origin[(r, c)] = top

    origin_text: dict[tuple[int, int], str] = {}
    style: dict[tuple[int, int], CellInfo] = {}
    bordered: set[tuple[int, int]] = set()

    for row in ws.iter_rows():
        for cell in row:
            pos = (cell.row, cell.column)
            # 罫線は被覆セルも含め全物理セルで収集(空の表セルを保護するため)
            if _has_border(cell):
                bordered.add(pos)
            if pos in covered_to_origin:
                continue  # 被覆セルは origin 側で扱う
            raw = cell.value
            if raw is None and options.formula_fallback and wsf is not None:
                vf = wsf.cell(row=cell.row, column=cell.column).value
                if isinstance(vf, str) and vf.startswith("="):
                    raw = vf
            text = _format_value(raw)
            if text.strip() != "":
                # ハイパーリンクは "テキスト (URL)" で URL を併記(両経路共通)
                link = getattr(cell, "hyperlink", None)
                target = getattr(link, "target", None) if link else None
                if target and target not in text:
                    text = f"{text} ({target})"
                origin_text[pos] = text
                style[pos] = _cell_style(cell)

    # 画像セル(ファイルは書かず、アンカー位置のみ収集)
    image_cells: dict[tuple[int, int], int] = {}
    for img in (getattr(ws, "_images", None) or []):
        anchor = _anchor_cell(img)
        if anchor is not None:
            image_cells[anchor] = image_cells.get(anchor, 0) + 1

    # 外接矩形(外周空白をトリム)
    bbox = _bbox(origin_text, span_of, image_cells)
    if bbox is None:
        # テキストが無くても罫線だけの表シートはありうる
        if not bordered:
            return None
        rs = [r for r, _ in bordered]
        cs = [c for _, c in bordered]
        bbox = (min(rs), min(cs), max(rs), max(cs))
    min_row, min_col, max_row, max_col = bbox

    # 罫線が「構造(表)」か「装飾(方眼紙の全面グリッド)」かを判定。
    # 表が主体のシートは罫線被覆率が高くなるため、装飾とみなすのは
    # 「ほぼ全セルが罫線」の極端な場合のみ(=保護を効かせる方向に倒す)。
    area = (max_row - min_row + 1) * (max_col - min_col + 1)
    in_box = sum(1 for (r, c) in bordered
                 if min_row <= r <= max_row and min_col <= c <= max_col)
    borders_are_structural = bool(bordered) and area > 0 and (in_box / area) < 0.9

    # 構造的な罫線なら、罫線つき空セル領域まで外接矩形を広げる
    if borders_are_structural:
        for (r, c) in bordered:
            min_row, min_col = min(min_row, r), min(min_col, c)
            max_row, max_col = max(max_row, r), max(max_col, c)

    body_font_size = _modal_font_size(style)

    return SheetModel(
        name=ws.title,
        min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col,
        origin_text=origin_text, span_of=span_of,
        covered_to_origin=covered_to_origin, style=style, image_cells=image_cells,
        bordered=bordered, borders_are_structural=borders_are_structural,
        body_font_size=body_font_size,
    )


def _has_border(cell) -> bool:
    border = cell.border
    return any(
        getattr(getattr(border, side), "style", None)
        for side in ("left", "right", "top", "bottom")
    )


def _modal_font_size(style: dict[tuple[int, int], CellInfo]) -> float | None:
    from collections import Counter
    sizes = [s.font_size for s in style.values() if s.font_size]
    if not sizes:
        return None
    return Counter(sizes).most_common(1)[0][0]


def _cell_style(cell) -> CellInfo:
    font = cell.font
    fill = cell.fill
    has_fill = bool(getattr(fill, "patternType", None)) and fill.patternType != "none"
    align = getattr(cell.alignment, "horizontal", None)
    return CellInfo(
        bold=bool(getattr(font, "bold", False)),
        font_size=getattr(font, "sz", None),
        has_fill=has_fill,
        has_border=_has_border(cell),
        h_align=align,
    )


def _bbox(origin_text, span_of, image_cells):
    min_r = min_c = max_r = max_c = None

    def expand(r1, c1, r2, c2):
        nonlocal min_r, min_c, max_r, max_c
        min_r = r1 if min_r is None else min(min_r, r1)
        min_c = c1 if min_c is None else min(min_c, c1)
        max_r = r2 if max_r is None else max(max_r, r2)
        max_c = c2 if max_c is None else max(max_c, c2)

    for (r, c) in origin_text:
        rs, cs = span_of.get((r, c), (1, 1))
        expand(r, c, r + rs - 1, c + cs - 1)
    for (r, c) in image_cells:
        expand(r, c, r, c)

    if min_r is None:
        return None
    return (min_r, min_c, max_r, max_c)
