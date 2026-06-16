"""Excel -> Markdown 変換のコア実装。

設計判断は要件定義書 v1.0 を参照:
- 1 シート = 原則 1 つの HTML <table>
- セル結合は colspan/rowspan に変換、被覆セルは出力しない
- 使用領域の外周空白のみ除去、内部の空セルは保持
- 文字値・結合・セル内改行のみ再現(色/太字/罫線は破棄)
- 数式はキャッシュ値、無ければ式文字列にフォールバック
- 画像は PNG/JPEG のみ抽出し、アンカー先 <td> にプレースホルダを差し込む
"""

from __future__ import annotations

import datetime as _dt
import logging
import re
from dataclasses import dataclass
from html import escape
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger("xlsx2md")

# 抽出対象とする画像フォーマット(マジックナンバー -> 拡張子)
_PNG_SIG = b"\x89PNG\r\n\x1a\n"
_JPEG_SIG = b"\xff\xd8\xff"


@dataclass
class Options:
    """変換オプション。"""

    image_dirname: str = "images"
    include_hidden: bool = True       # 隠し行/列を出力に含めるか
    formula_fallback: bool = True     # 数式キャッシュが無い時に式文字列を使うか


# --------------------------------------------------------------------------- #
# 公開 API
# --------------------------------------------------------------------------- #
def convert(xlsx_path: str | Path, out_dir: str | Path,
            options: Options | None = None) -> list[Path]:
    """xlsx を変換し、生成した .md ファイルのパス一覧を返す。

    出力構成:
        {out_dir}/{xlsxファイル名}/{連番}_{シート名}.md
        {out_dir}/{xlsxファイル名}/images/...
    """
    options = options or Options()
    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)

    # data_only: 数式のキャッシュ値を取得。fallback 用に式文字列版も開く。
    wb_data = load_workbook(xlsx_path, data_only=True)
    wb_form = (load_workbook(xlsx_path, data_only=False)
               if options.formula_fallback else None)

    doc_dir = out_dir / _sanitize(xlsx_path.stem)
    images_dir = doc_dir / options.image_dirname
    doc_dir.mkdir(parents=True, exist_ok=True)

    produced: list[Path] = []
    for idx, sheet_name in enumerate(wb_data.sheetnames, start=1):
        ws = wb_data[sheet_name]
        wsf = wb_form[sheet_name] if wb_form is not None else None

        image_map = _extract_images(ws, images_dir, _sanitize(sheet_name), options)
        body = render_faithful_table(ws, wsf, image_map, options)

        md = f"# {sheet_name}\n\n{body}\n"
        fname = f"{idx:02d}_{_sanitize(sheet_name)}.md"
        out_path = doc_dir / fname
        out_path.write_text(md, encoding="utf-8")
        produced.append(out_path)
        logger.info("wrote %s", out_path)

    return produced


# --------------------------------------------------------------------------- #
# シート -> HTML テーブル
# --------------------------------------------------------------------------- #
def render_faithful_table(ws, wsf, image_map: dict[tuple[int, int], list[str]],
                          options: Options) -> str:
    """シート全体を 1 つの忠実な HTML テーブルに描く(v1 方式)。

    v2 ではトップレベルの出力ではなく、(a) 分割が外れた時の突き合わせ用ダンプ、
    (b) 将来の LLMInterpreter に渡す忠実な中間表現、として温存する。
    結合→colspan/rowspan の変換ロジックは ⑤ の表ブロック描画でも再利用予定。
    """
    merges = list(ws.merged_cells.ranges)

    # 結合: 左上セル -> (rowspan, colspan)、被覆セル集合
    span_of: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for mr in merges:
        top = (mr.min_row, mr.min_col)
        span_of[top] = (mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != top:
                    covered.add((r, c))

    bbox = _content_bbox(ws, wsf, image_map, merges, options)
    if bbox is None:
        return "_(空のシート)_"
    min_row, min_col, max_row, max_col = bbox

    # 非表示行/列(include_hidden=False のとき除外)
    hidden_rows, hidden_cols = _hidden_sets(ws, options)

    lines = ["<table>"]
    for r in range(min_row, max_row + 1):
        if r in hidden_rows:
            continue
        cells_html: list[str] = []
        for c in range(min_col, max_col + 1):
            if c in hidden_cols:
                continue
            if (r, c) in covered:
                continue  # 結合に被覆されたセルは出力しない
            attrs = ""
            if (r, c) in span_of:
                rowspan, colspan = span_of[(r, c)]
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
            content = _cell_content(ws, wsf, r, c, image_map, options)
            cells_html.append(f"<td{attrs}>{content}</td>")
        lines.append("<tr>" + "".join(cells_html) + "</tr>")
    lines.append("</table>")
    return "\n".join(lines)


def _cell_content(ws, wsf, r: int, c: int,
                  image_map: dict[tuple[int, int], list[str]],
                  options: Options) -> str:
    cell = ws.cell(row=r, column=c)
    raw = cell.value
    if raw is None and options.formula_fallback and wsf is not None:
        vf = wsf.cell(row=r, column=c).value
        if isinstance(vf, str) and vf.startswith("="):
            raw = vf  # キャッシュ無し -> 式文字列にフォールバック

    text = _format_value(raw)
    html = escape(text).replace("\n", "<br>")

    # ハイパーリンク: "テキスト (URL)" 形式で URL を併記
    link = getattr(cell, "hyperlink", None)
    target = getattr(link, "target", None) if link else None
    if target and target not in text:
        sep = " " if html else ""
        html = f"{html}{sep}({escape(target)})"

    # アンカーされた画像のプレースホルダを差し込む
    for fname in image_map.get((r, c), []):
        col_letter = get_column_letter(c)
        placeholder = f"[画像: {fname} | アンカー {col_letter}{r}]"
        html = (html + "<br>" + escape(placeholder)) if html else escape(placeholder)

    return html


def _format_value(raw) -> str:
    """セル値を表示用文字列に整形する。

    最小版: 日付/時刻は読みやすい固定書式、真偽値は TRUE/FALSE、
    それ以外は str()。Excel の number_format 完全適用は将来課題。
    """
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "TRUE" if raw else "FALSE"
    if isinstance(raw, _dt.datetime):
        if raw.time() == _dt.time(0, 0):
            return raw.strftime("%Y/%m/%d")
        return raw.strftime("%Y/%m/%d %H:%M:%S")
    if isinstance(raw, _dt.date):
        return raw.strftime("%Y/%m/%d")
    if isinstance(raw, _dt.time):
        return raw.strftime("%H:%M:%S")
    return str(raw)


# --------------------------------------------------------------------------- #
# 画像抽出
# --------------------------------------------------------------------------- #
def _extract_images(ws, images_dir: Path, sheet_slug: str,
                    options: Options) -> dict[tuple[int, int], list[str]]:
    """シート内の埋め込み画像(PNG/JPEG)を抽出し、(row, col)->[ファイル名] を返す。"""
    image_map: dict[tuple[int, int], list[str]] = {}
    images = getattr(ws, "_images", None) or []

    for i, img in enumerate(images, start=1):
        raw = _image_bytes(img)
        if raw is None:
            logger.warning("画像データを取得できませんでした (sheet=%s, idx=%d)",
                           ws.title, i)
            continue
        ext = _detect_ext(raw)
        if ext is None:
            logger.warning("非対応の画像形式をスキップ (sheet=%s, idx=%d)",
                           ws.title, i)
            continue

        anchor_cell = _anchor_cell(img)
        if anchor_cell is None:
            logger.warning("画像のアンカー位置を取得できませんでした (sheet=%s, idx=%d)",
                           ws.title, i)
            continue

        images_dir.mkdir(parents=True, exist_ok=True)
        fname = f"{sheet_slug}_img{i}.{ext}"
        (images_dir / fname).write_bytes(raw)
        image_map.setdefault(anchor_cell, []).append(fname)

    return image_map


def _image_bytes(img) -> bytes | None:
    """openpyxl の Image オブジェクトからバイト列を取り出す(版差を吸収)。"""
    data = getattr(img, "_data", None)
    if callable(data):
        try:
            return data()
        except Exception:  # noqa: BLE001
            pass
    ref = getattr(img, "ref", None)
    if ref is not None:
        if hasattr(ref, "getvalue"):
            try:
                return ref.getvalue()
            except Exception:  # noqa: BLE001
                pass
        if hasattr(ref, "read"):
            try:
                ref.seek(0)
                return ref.read()
            except Exception:  # noqa: BLE001
                pass
    return None


def _detect_ext(raw: bytes) -> str | None:
    if raw.startswith(_PNG_SIG):
        return "png"
    if raw.startswith(_JPEG_SIG):
        return "jpg"
    return None


def _anchor_cell(img) -> tuple[int, int] | None:
    """画像のアンカー先セルを (row, col) (1-indexed) で返す。"""
    anchor = getattr(img, "anchor", None)
    frm = getattr(anchor, "_from", None)
    if frm is not None and hasattr(frm, "row") and hasattr(frm, "col"):
        return (frm.row + 1, frm.col + 1)  # openpyxl は 0-indexed
    # 文字列アンカー("A1" 等)のフォールバック
    if isinstance(anchor, str):
        m = re.match(r"([A-Za-z]+)(\d+)", anchor)
        if m:
            from openpyxl.utils import column_index_from_string
            return (int(m.group(2)), column_index_from_string(m.group(1)))
    return None


# --------------------------------------------------------------------------- #
# 補助
# --------------------------------------------------------------------------- #
def _content_bbox(ws, wsf, image_map, merges, options
                  ) -> tuple[int, int, int, int] | None:
    """内容(値あり・結合・画像)の外接矩形を返す。外周空白のトリミング用。"""
    min_r = min_c = None
    max_r = max_c = None

    def expand(r1, c1, r2, c2):
        nonlocal min_r, min_c, max_r, max_c
        min_r = r1 if min_r is None else min(min_r, r1)
        min_c = c1 if min_c is None else min(min_c, c1)
        max_r = r2 if max_r is None else max(max_r, r2)
        max_c = c2 if max_c is None else max(max_c, c2)

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None and options.formula_fallback and wsf is not None:
                vf = wsf.cell(row=cell.row, column=cell.column).value
                if isinstance(vf, str) and vf.startswith("="):
                    v = vf
            if v is not None and str(v).strip() != "":
                expand(cell.row, cell.column, cell.row, cell.column)

    for mr in merges:
        top = ws.cell(row=mr.min_row, column=mr.min_col).value
        if top is not None and str(top).strip() != "":
            expand(mr.min_row, mr.min_col, mr.max_row, mr.max_col)

    for (r, c) in image_map:
        expand(r, c, r, c)

    if min_r is None:
        return None
    return (min_r, min_c, max_r, max_c)


def _hidden_sets(ws, options: Options) -> tuple[set[int], set[int]]:
    if options.include_hidden:
        return set(), set()
    hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
    hidden_cols = set()
    for letter, dim in ws.column_dimensions.items():
        if dim.hidden:
            from openpyxl.utils import column_index_from_string
            hidden_cols.add(column_index_from_string(letter))
    return hidden_rows, hidden_cols


_SANITIZE_RE = re.compile(r'[\\/:*?"<>|\s]+')


def _sanitize(name: str) -> str:
    """ファイル名に使えない文字を _ に置換する。"""
    cleaned = _SANITIZE_RE.sub("_", name.strip())
    return cleaned.strip("_") or "sheet"
